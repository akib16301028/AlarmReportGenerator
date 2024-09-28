# alarm_report_generator_streamlit.py

import pandas as pd
import os
import streamlit as st
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

# Set the page configuration
st.set_page_config(
    page_title="Alarm Report Generator",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Title of the app
st.title("📊 Alarm Report Generator")

# Description
st.markdown("""
This application processes your Excel files to generate formatted alarm summary reports. 
Upload your Excel file, and the app will handle the rest!
""")

# Sidebar for file upload
st.sidebar.header("Upload Your Excel File")
uploaded_file = st.sidebar.file_uploader(
    "Choose an Excel file", type=["xlsx"], accept_multiple_files=False
)

# Function to process the Excel file
def process_file(file):
    # Read the Excel file, assuming header is on the 3rd row (index=2)
    df = pd.read_excel(file, header=2)

    # Step 0: Identify leased sites (sites starting with 'L' in column 'Site')
    if 'Site' in df.columns:
        leased_sites = df[df['Site'].str.startswith('L', na=False)]
        leased_site_names = leased_sites['Site'].tolist()
    else:
        st.error(f"'Site' column not found in the data! Columns available: {df.columns.tolist()}")
        st.stop()

    # Step 1: Identify BANJO and Non BANJO sites in the 'Site Alias' column (Column 'Site Alias')
    if 'Site Alias' in df.columns:
        df['Site Type'] = df['Site Alias'].apply(lambda x: 'BANJO' if '(BANJO)' in str(x) else 'Non BANJO')
        # Extract client information (BL, GP, Robi) from Site Alias
        df['Client'] = df['Site Alias'].apply(
            lambda x: 'BL' if '(BL)' in str(x)
            else 'GP' if '(GP)' in str(x)
            else 'Robi' if '(ROBI)' in str(x)
            else 'BANJO'
        )
    else:
        st.error(f"'Site Alias' column not found in the data! Columns available: {df.columns.tolist()}")
        st.stop()

    # Step 2: Filter relevant alarm categories in the 'Alarm Name' column
    alarm_categories = ['Battery Low', 'Mains Fail', 'DCDB-01 Primary Disconnect', 'PG Run']
    if 'Alarm Name' in df.columns:
        filtered_df = df[df['Alarm Name'].isin(alarm_categories)]
    else:
        st.error(f"'Alarm Name' column not found in the data! Columns available: {df.columns.tolist()}")
        st.stop()

    # Step 3: Exclude leased sites for "DCDB-01 Primary Disconnect" alarm
    dcdb_df = filtered_df[filtered_df['Alarm Name'] == 'DCDB-01 Primary Disconnect']
    non_leased_dcdb_df = dcdb_df[~dcdb_df['Site'].isin(leased_site_names)]  # Exclude leased sites

    # Step 4: Create a combined DataFrame with all alarms, replacing the filtered "DCDB-01 Primary Disconnect"
    filtered_df = pd.concat([
        filtered_df[filtered_df['Alarm Name'] != 'DCDB-01 Primary Disconnect'], 
        non_leased_dcdb_df
    ])

    # Step 5: Generate a client-wise table for tenant counts for each alarm
    if 'Tenant' in df.columns:
        client_table = df.groupby(['Alarm Name', 'Client', 'Tenant']).size().reset_index(name='Count')

        # Pivot the table to get separate columns for each Tenant under each Alarm and Client
        pivot_table = client_table.pivot_table(
            index=['Alarm Name', 'Client'], 
            columns='Tenant', 
            values='Count',
            fill_value=0
        )
        pivot_table.reset_index(inplace=True)
        pivot_table.columns.name = None
    else:
        st.error(f"'Tenant' column not found in the data! Columns available: {df.columns.tolist()}")
        st.stop()

    # Check for RIO (Cluster) and Subcenter (Zone)
    if 'Cluster' in df.columns and 'Zone' in df.columns:
        # Step 7: Group data by Alarm Name, Client (BANJO, GP, BL, Robi), RIO, Subcenter, and count occurrences
        summary_table = filtered_df.groupby(['Alarm Name', 'Cluster', 'Zone', 'Client']).size().reset_index(name='Alarm Count')

        # Renaming for better readability
        summary_table = summary_table.rename(columns={
            'Cluster': 'RIO',
            'Zone': 'Subcenter'
        })

        # Sort the summary_table by RIO and Alarm Count (descending)
        summary_table = summary_table.sort_values(by=['RIO', 'Alarm Count'], ascending=[True, False])

        # Define color fill for each RIO and header
        rios = summary_table['RIO'].unique()
        colors = ['FFFF99', 'FF9999', 'D5A6F2', '99CCFF', 'FFCC99']  # Example color codes
        color_map = {rio: PatternFill(start_color=color, end_color=color, fill_type="solid") 
                     for rio, color in zip(rios, colors)}

        # Define color fill and font for headers and total row
        header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        total_row_fill = PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")
        overall_row_fill = PatternFill(start_color="E0EBEB", end_color="E0EBEB", fill_type="solid")
        bold_font = Font(bold=True)
        border_style = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Create a new Workbook
        wb = Workbook()

        # Save the original DataFrame to a new sheet
        ws_original = wb.active
        ws_original.title = "Original Data"
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_original.cell(row=r_idx, column=c_idx, value=value)

        # Add each alarm-specific table to a separate sheet
        for alarm in alarm_categories:
            alarm_data = summary_table[summary_table['Alarm Name'] == alarm]

            if alarm_data.empty:
                continue  # Skip if there's no data for the alarm

            # Pivot the table to get separate columns for each client (BANJO, GP, BL, Robi)
            pivot_table = alarm_data.pivot_table(
                index=['RIO', 'Subcenter'], 
                columns='Client', 
                values='Alarm Count',
                fill_value=0
            )
            pivot_table.reset_index(inplace=True)
            pivot_table.columns.name = None

            # Sort the pivot table data first by RIO, then by counts in descending order
            pivot_table = pivot_table.sort_values(
                by=['RIO'] + list(pivot_table.columns[2:]),
                ascending=[True] + [False] * (len(pivot_table.columns) - 2)
            )

            # Create a new sheet for the alarm
            ws = wb.create_sheet(title=alarm)

            # Write the alarm name in bold
            alarm_cell = ws.cell(row=1, column=1, value=alarm)
            alarm_cell.font = Font(bold=True)
            alarm_cell.alignment = center_alignment

            # Write the pivot table data
            for r_idx, row in enumerate(dataframe_to_rows(pivot_table, index=False, header=True), 2):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = center_alignment

                    # Apply color fill based on RIO
                    rio = row[pivot_table.columns.get_loc('RIO')]
                    if rio in color_map:
                        cell.fill = color_map[rio]

                    # Apply border style
                    cell.border = border_style

            # Merge cells for identical RIO values
            rio_col_idx = 1  # Index of RIO column in the pivot table
            current_rio = None
            start_row = None
            for r_idx in range(2, len(pivot_table) + 2):
                cell_value = ws.cell(row=r_idx, column=rio_col_idx).value
                if cell_value == current_rio:
                    continue
                if current_rio is not None and start_row is not None:
                    ws.merge_cells(
                        start_row=start_row, 
                        start_column=rio_col_idx, 
                        end_row=r_idx - 1, 
                        end_column=rio_col_idx
                    )
                current_rio = cell_value
                start_row = r_idx
            if start_row is not None:
                ws.merge_cells(
                    start_row=start_row, 
                    start_column=rio_col_idx, 
                    end_row=len(pivot_table) + 1,
                    end_column=rio_col_idx
                )

            # Format header cells
            for c_idx in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=2, column=c_idx)
                header_cell.font = bold_font
                header_cell.alignment = center_alignment
                header_cell.fill = header_fill

            # Add 'Total' row
            total_row_index = len(pivot_table) + 3
            ws.cell(row=total_row_index, column=1, value='Total').font = bold_font
            ws.cell(row=total_row_index, column=2, value='')  # Empty column B
            ws.cell(row=total_row_index, column=1).alignment = center_alignment
            ws.cell(row=total_row_index, column=2).alignment = center_alignment

            for col_idx in range(3, ws.max_column + 1):
                total = sum(
                    ws.cell(row=r_idx, column=col_idx).value or 0 
                    for r_idx in range(3, len(pivot_table) + 3)
                )
                ws.cell(row=total_row_index, column=col_idx, value=total)
                ws.cell(row=total_row_index, column=col_idx).font = bold_font
                ws.cell(row=total_row_index, column=col_idx).alignment = center_alignment
                ws.cell(row=total_row_index, column=col_idx).fill = total_row_fill
                ws.cell(row=total_row_index, column=col_idx).border = border_style

            # Add 'Overall' row
            overall_row_index = total_row_index + 1
            ws.cell(row=overall_row_index, column=1, value='Overall').font = bold_font
            ws.cell(row=overall_row_index, column=1).alignment = center_alignment
            ws.cell(row=overall_row_index, column=1).fill = overall_row_fill
            ws.cell(row=overall_row_index, column=1).border = border_style

            # Calculate overall total from the 'Total' row
            overall_total = 0
            for col_idx in range(3, ws.max_column + 1):
                total_value = ws.cell(row=total_row_index, column=col_idx).value or 0
                overall_total += total_value

            # Place the overall total in column B
            ws.cell(row=overall_row_index, column=2, value=overall_total)  # Place total in column B
            ws.cell(row=overall_row_index, column=2).font = bold_font
            ws.cell(row=overall_row_index, column=2).alignment = center_alignment
            ws.cell(row=overall_row_index, column=2).fill = overall_row_fill
            ws.cell(row=overall_row_index, column=2).border = border_style

            # Clear the other cells in the 'Overall' row if they were populated
            for col_idx in range(3, ws.max_column + 1):
                ws.cell(row=overall_row_index, column=col_idx, value='')

            # Define a function to apply borders to all cells in the worksheet
            def apply_borders(ws):
                thin_border = Border(
                    left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin")
                )
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = thin_border

            # Apply borders to all cells in the worksheet
            apply_borders(ws)

            # Adjust column widths to fit the data
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width

        # Save the workbook to a BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output
    else:
        st.error(f"'Cluster' and/or 'Zone' column(s) not found in the data! Columns available: {df.columns.tolist()}")
        st.stop()

# Main logic
if uploaded_file is not None:
    with st.spinner('Processing...'):
        processed_file = process_file(uploaded_file)
    st.success('Report generated successfully!')

    # Provide a download button
    st.download_button(
        label="📥 Download Report",
        data=processed_file,
        file_name="Alarm_Summary_Report_Formatted_By_Alarm.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Optionally, display the original data and summary tables
    if st.checkbox("Show Original Data"):
        df = pd.read_excel(uploaded_file, header=2)
        st.dataframe(df)

    if st.checkbox("Show Summary Table"):
        # Re-process to get summary_table for display
        df = pd.read_excel(uploaded_file, header=2)
        # Repeat steps to generate summary_table without formatting
        if 'Site' in df.columns and 'Site Alias' in df.columns and 'Alarm Name' in df.columns and 'Tenant' in df.columns and 'Cluster' in df.columns and 'Zone' in df.columns:
            leased_sites = df[df['Site'].str.startswith('L', na=False)]
            leased_site_names = leased_sites['Site'].tolist()

            df['Site Type'] = df['Site Alias'].apply(lambda x: 'BANJO' if '(BANJO)' in str(x) else 'Non BANJO')
            df['Client'] = df['Site Alias'].apply(
                lambda x: 'BL' if '(BL)' in str(x)
                else 'GP' if '(GP)' in str(x)
                else 'Robi' if '(ROBI)' in str(x)
                else 'BANJO'
            )

            alarm_categories = ['Battery Low', 'Mains Fail', 'DCDB-01 Primary Disconnect', 'PG Run']
            filtered_df = df[df['Alarm Name'].isin(alarm_categories)]

            dcdb_df = filtered_df[filtered_df['Alarm Name'] == 'DCDB-01 Primary Disconnect']
            non_leased_dcdb_df = dcdb_df[~dcdb_df['Site'].isin(leased_site_names)]
            filtered_df = pd.concat([
                filtered_df[filtered_df['Alarm Name'] != 'DCDB-01 Primary Disconnect'], 
                non_leased_dcdb_df
            ])

            client_table = df.groupby(['Alarm Name', 'Client', 'Tenant']).size().reset_index(name='Count')
            pivot_table = client_table.pivot_table(
                index=['Alarm Name', 'Client'], 
                columns='Tenant', 
                values='Count',
                fill_value=0
            )
            pivot_table.reset_index(inplace=True)
            pivot_table.columns.name = None

            summary_table = filtered_df.groupby(['Alarm Name', 'Cluster', 'Zone', 'Client']).size().reset_index(name='Alarm Count')
            summary_table = summary_table.rename(columns={'Cluster': 'RIO', 'Zone': 'Subcenter'})
            summary_table = summary_table.sort_values(by=['RIO', 'Alarm Count'], ascending=[True, False])

            st.dataframe(summary_table)
        else:
            st.error("Required columns are missing in the uploaded file.")
else:
    st.info("Please upload an Excel file to get started.")
